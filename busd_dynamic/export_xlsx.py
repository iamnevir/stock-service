"""
export_xlsx.py
==============
Export kết quả backtest từ collection busd_dynamic ra file Excel.

Mỗi dòng = 1 (week, src, mode) có đủ play_strategy + play_report.

Cột đầu ra:
    week            | tuần hiện tại (vd: 2026-W02)
    next_week       | tuần tiếp theo backtest chạy trên đó
    week_start      | ngày đầu tuần (từ lst_day)
    week_end        | ngày cuối tuần (từ lst_day)
    src             | nguồn dữ liệu (hose500, vn30, ...)
    mode            | loại MA (ema, sma, ...)
    play_size       | số lượng strategies trong play_strategy
    play_sharpe     | play_mean_sharpe (sharpe trung bình IS)
    play_tvr        | play_mean_tvr (tvr trung bình IS)
    play_corr       | play_correlation (correlation trung bình giữa strategies)
    sharpe          | từ play_report (kết quả OS next_week)
    tvr             | từ play_report
    total_net_profit| từ play_report
    mdd_percent     | từ play_report
    profit_percent  | từ play_report

Chạy:
    python -m busd_dynamic.export_xlsx
    python -m busd_dynamic.export_xlsx --out /tmp/result.xlsx
    python -m busd_dynamic.export_xlsx --src hose500 --mode ema
"""

import argparse
from typing import List, Optional

import pandas as pd
from pymongo import MongoClient

from busd_auto.utils import get_mongo_uri
from upload import upload_file_to_drive

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _build_next_week_map(collection) -> dict:
    """Build dict {week -> next_week} từ toàn bộ tuần trong collection."""
    all_weeks = sorted(
        doc["week"]
        for doc in collection.find({"week": {"$exists": True}}, {"week": 1})
        if doc.get("week")
    )
    week_map = {}
    for i, w in enumerate(all_weeks):
        week_map[w] = all_weeks[i + 1] if i + 1 < len(all_weeks) else None
    print(f"[export] Đã build next_week map: {len(all_weeks)} tuần")
    return week_map


def _fetch_rows(
    collection,
    next_week_map: dict,
    src_filter: Optional[str],
    mode_filter: Optional[str],
    week_filter: Optional[str],
    require_report: bool,
) -> List[dict]:
    """Lấy toàn bộ rows từ DB và flatten thành list of dict."""

    # Query docs có play_strategy
    query: dict = {
        "srcs": {
            "$elemMatch": {
                "play_strategy": {"$exists": True, "$not": {"$size": 0}},
            }
        }
    }
    if week_filter:
        query["week"] = week_filter

    projection = {"week": 1, "lst_day": 1, "srcs": 1}
    docs = list(collection.find(query, projection).sort("week", 1))
    print(f"[export] Tìm thấy {len(docs)} doc(s) có play_strategy")

    rows = []
    for doc in docs:
        week = doc.get("week")
        if not week:
            continue

        lst_day = sorted(int(d) for d in doc.get("lst_day", []))
        week_start = lst_day[0] if lst_day else None
        week_end = lst_day[-1] if lst_day else None
        next_week = doc.get("next_week") or next_week_map.get(week)

        for src_item in doc.get("srcs", []):
            src = src_item.get("src")
            mode = src_item.get("mode")
            play_strategy: list = src_item.get("play_strategy", [])

            # Filters
            if not play_strategy:
                continue
            if src_filter and src != src_filter:
                continue
            if mode_filter and mode != mode_filter:
                continue

            play_report: dict = src_item.get("play_report", {}) or {}

            if require_report and not play_report:
                continue

            row = {
                "strategies": "\n".join(play_strategy),
                "week": week,
                "week_start": week_start,
                "week_end": week_end,
                "src": src,
                "mode": mode,
                # Play strategy metadata (IS)
                "play_size": len(play_strategy),
                "play_sharpe": src_item.get("play_mean_sharpe"),
                "play_tvr": src_item.get("play_mean_tvr"),
                "play_corr": src_item.get("play_correlation"),
                # Play report (OS = next_week)
                "sharpe": play_report.get("sharpe"),
                "tvr": play_report.get("tvr"),
                "total_net_profit": play_report.get("total_net_profit"),
                "mdd_percent": play_report.get("mdd_percent"),
                "profit_percent": play_report.get("profit_percent"),
            }
            rows.append(row)

    return rows


# Mapping key -> display name
_COLUMN_NAMES = {
    "strategies":        "Strategies (50)",
    "week":              "Tuần IS",
    "week_start":        "Ngày bắt đầu",
    "week_end":          "Ngày kết thúc",
    "src":               "Nguồn",
    "mode":              "Mode MA",
    "play_size":         "Số strategy",
    "play_sharpe":       "Sharpe IS TB",
    "play_tvr":          "TVR IS TB",
    "play_corr":         "Tương quan TB",
    "sharpe":            "Sharpe OS",
    "tvr":               "TVR OS",
    "total_net_profit":  "Net Profit OS",
    "mdd_percent":       "MDD % OS",
    "profit_percent":    "Profit % OS",
}

_COLUMN_ORDER = list(_COLUMN_NAMES.keys())


def _build_dataframe(rows: List[dict]) -> pd.DataFrame:
    """Chuyển rows thành DataFrame với thứ tự cột cố định và tên đẹp."""
    df = pd.DataFrame(rows, columns=_COLUMN_ORDER)
    df = df.rename(columns=_COLUMN_NAMES)
    return df


def _export_excel(df: pd.DataFrame, out_path: str) -> None:
    """Ghi DataFrame ra file Excel với format đẹp."""
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="backtest_results")

        ws = writer.sheets["backtest_results"]
        wb = writer.book

        # ---- Import styles ----
        from openpyxl.styles import (
            Alignment, Font, PatternFill, Border, Side, numbers
        )
        from openpyxl.utils import get_column_letter

        HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
        HEADER_FONT  = Font(color="FFFFFF", bold=True, size=10)
        CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=False)
        LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        STRAT_ALIGN  = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
        THIN_BORDER  = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        IS_META_FILL = PatternFill("solid", fgColor="D9E1F2")  # xanh nhạt → IS metadata
        OS_FILL      = PatternFill("solid", fgColor="E2EFDA")  # xanh lá nhạt → OS results
        STRAT_FILL   = PatternFill("solid", fgColor="FFF2CC")  # vàng nhạt → Strategies
        ROW_ALT_FILL = PatternFill("solid", fgColor="F5F5F5")

        # col index (1-based) based on _COLUMN_ORDER:
        # 1=strategies, 2=week, 3=week_start, 4=week_end, 5=src, 6=mode,
        # 7=play_size, 8=play_sharpe, 9=play_tvr, 10=play_corr,
        # 11=sharpe, 12=tvr, 13=total_net_profit, 14=mdd_percent, 15=profit_percent
        IS_META_COLS = {7, 8, 9, 10}
        OS_COLS      = {11, 12, 13, 14, 15}

        num_fmt_2 = "0.00"
        num_fmt_3 = "0.000"
        num_fmt_0 = "#,##0"

        col_formats = {
            7:  num_fmt_0,  # play_size
            8:  num_fmt_3,  # play_sharpe
            9:  num_fmt_3,  # play_tvr
            10: num_fmt_3,  # play_corr
            11: num_fmt_3,  # sharpe
            12: num_fmt_3,  # tvr
            13: num_fmt_2,  # total_net_profit
            14: num_fmt_2,  # mdd_percent
            15: num_fmt_2,  # profit_percent
        }

        # Header row
        for cell in ws[1]:
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.alignment = CENTER
            cell.border    = THIN_BORDER

        # Data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            is_alt = (row_idx % 2 == 0)
            for cell in row:
                col_i = cell.column
                cell.border = THIN_BORDER

                if col_i == 1:  # Strategies
                    cell.fill      = STRAT_FILL
                    cell.alignment = STRAT_ALIGN
                elif col_i in IS_META_COLS:
                    cell.fill      = IS_META_FILL
                    cell.alignment = CENTER
                elif col_i in OS_COLS:
                    cell.fill      = OS_FILL
                    cell.alignment = CENTER
                else:
                    cell.alignment = LEFT
                    if is_alt:
                        cell.fill = ROW_ALT_FILL

                if col_i in col_formats:
                    cell.number_format = col_formats[col_i]

                # Colour OS Sharpe: green > 0, red < 0
                if col_i == 11 and cell.value is not None:
                    try:
                        val = float(cell.value)
                        if val > 0:
                            cell.font = Font(color="375623", bold=True, size=10)
                        elif val < 0:
                            cell.font = Font(color="C00000", bold=True, size=10)
                    except (TypeError, ValueError):
                        pass

        # Column widths
        col_widths = {
            1:  28,  # Strategies
            2:  12,  # Tuần IS
            3:  14,  # Ngày bắt đầu
            4:  14,  # Ngày kết thúc
            5:  14,  # Nguồn
            6:   8,  # Mode MA
            7:  11,  # Số strategy
            8:  14,  # Sharpe IS TB
            9:  11,  # TVR IS TB
            10: 14,  # Tương quan TB
            11: 11,  # Sharpe OS
            12:  9,  # TVR OS
            13: 16,  # Net Profit OS
            14: 12,  # MDD % OS
            15: 14,  # Profit % OS
        }
        for col_i, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col_i)].width = width

        # Row heights: strategies column needs taller rows
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 52

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto-filter
        ws.auto_filter.ref = ws.dimensions

    print(f"[export] ✅ Đã xuất {len(df)} dòng → {out_path}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Export kết quả play_report từ busd_dynamic ra file Excel"
    )
    parser.add_argument(
        "--out", type=str,
        default="/home/ubuntu/nevir/busd_dynamic/backtest_results.xlsx",
        help="Đường dẫn file output (.xlsx)"
    )
    parser.add_argument("--week", type=str, default=None, help="Chỉ export 1 tuần, vd: 2026-W02")
    parser.add_argument("--src",  type=str, default=None, help="Filter theo src, vd: hose500")
    parser.add_argument("--mode", type=str, default=None, help="Filter theo mode, vd: ema")
    parser.add_argument(
        "--all", action="store_true",
        help="Export cả những dòng chưa có play_report (để review)"
    )
    args = parser.parse_args()

    mongo_client = MongoClient(get_mongo_uri("mgc3"))
    collection = mongo_client["busd"]["busd_dynamic"]

    next_week_map = _build_next_week_map(collection)

    rows = _fetch_rows(
        collection=collection,
        next_week_map=next_week_map,
        src_filter=args.src,
        mode_filter=args.mode,
        week_filter=args.week,
        require_report=not args.all,
    )

    mongo_client.close()

    if not rows:
        print("[export] Không có dữ liệu để export.")
        return

    df = _build_dataframe(rows)
    print(f"[export] Tổng rows: {len(df)}")
    # print(df.to_string(index=False))

    _export_excel(df, args.out)
    upload_file_to_drive("/home/ubuntu/nevir/busd_dynamic/backtest_results.xlsx")

if __name__ == "__main__":
    main()
