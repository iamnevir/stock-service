import json
import numpy as np
from itertools import product
import pandas as pd
import pickle
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor, as_completed
from multiprocessing import cpu_count
import time
from core import Simulator
from alpha_func_lib import Domains
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule
import os
from datetime import datetime

from googleapiclient.errors import HttpError
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials

class PKLToXLSXExporter:
    def __init__(self, df, gen=None):
        """
        Initialize exporter với file PKL
        
        Args:
            pkl_file_path (str): Đường dẫn tới file PKL
        """
        self.df = df
        self.column_mapping = {
            # Mapping từ tên cột gốc sang tên cột mới
            'alphaName': 'Strategy',
            'freq': 'Frequency',
            'fee': 'Fee',
            
            'sharpe': 'Sharpe Ratio',
            "hhi": "HHI",
            'daily_sharpe': 'Daily Sharpe',
            "psr": "PSR (%)",
            "dsr": "DSR (%)",
            "E": "E", 
            'mdd': 'MDD',
            'mddPct': 'MDD (%)',
            'ppc': 'PPC',
            'tvr': 'TVR',
            "netProfit": 'Net Profit',
            "overnight_pnl": "Overnight PnL",
            'start': 'Start',
            'end': 'End',
            'total_return': 'Total Return (%)',
            'annual_return': 'Annual Return (%)',
            'volatility': 'Volatility (%)',
            'win_rate': 'Win Rate (%)',
            'profit_factor': 'Profit Factor',
            'total_trades': 'Total Trades',
            'avg_trade_return': 'Avg Trade Return (%)',
            'best_trade': 'Best Trade (%)',
            'worst_trade': 'Worst Trade (%)',
            'consecutive_wins': 'Max Consecutive Wins',
            'consecutive_losses': 'Max Consecutive Losses',
            'calmar_ratio': 'Calmar Ratio',
            'sortino_ratio': 'Sortino Ratio',
            'params': 'Parameters'
        }
        if gen == "1_1":
            self.column_mapping['threshold'] = 'Threshold'
            self.column_mapping['halflife'] = 'Halflife'
        elif gen == "1_2":
            self.column_mapping['upper'] = 'Upper'
            self.column_mapping['lower'] = 'Lower'
        elif gen == "1_3":
            self.column_mapping['score'] = 'Score'
            self.column_mapping['entry'] = 'Entry'
            self.column_mapping['exit'] = 'Exit'
        elif gen == "1_4":
            self.column_mapping['entry'] = 'Entry'
            self.column_mapping['exit'] = 'Exit'
            self.column_mapping['smooth'] = 'Smooth'
    
    def format_dataframe(self):
        """
        Format và rename columns, sort data
        """
        if self.df is None:
            print("No data loaded. Please run load_data() first.")
            return False
        
        # Tìm tất cả columns có trong data
        available_columns = list(self.df.columns)
        print(f"Available columns: {available_columns}")
        
        # Tách columns thành các nhóm
        mapped_columns = []  # Columns có trong mapping
        param_columns = []   # Columns là parameters
        other_columns = []   # Columns khác
        
        for col in available_columns:
            if col in self.column_mapping:
                mapped_columns.append(col)
            elif col.startswith('param_') or col in ['window', 'factor']:
                # Các tên parameter phổ biến
                param_columns.append(col)
            else:
                other_columns.append(col)
        
        print(f"Mapped columns: {mapped_columns}")
        print(f"Parameter columns: {param_columns}")
        print(f"Other columns: {other_columns}")
        
        # Chọn columns để giữ lại (mapped + param + some others)
        columns_to_keep = mapped_columns + param_columns
        
        # Thêm các columns khác quan trọng nếu có
        important_other_cols = ['params']  # Giữ lại params gốc nếu có
        for col in important_other_cols:
            if col in other_columns:
                columns_to_keep.append(col)
        
        # Chỉ lấy columns có trong data
        self.df = self.df[columns_to_keep].copy()
        for drop_col in ['fee', 'start', 'end']:
            if drop_col in self.df.columns:
                self.df.drop(columns=[drop_col], inplace=True)
        # Rename các columns có mapping
        rename_dict = {k: v for k, v in self.column_mapping.items() if k in self.df.columns}
        self.df.rename(columns=rename_dict, inplace=True)
        
        # Rename parameter columns để dễ đọc hơn
        param_rename_dict = {}
        for col in param_columns:
            if col.startswith('param_'):
                new_name = f"Param: {col[6:].title()}"  # Remove 'param_' prefix và capitalize
            else:
                new_name = f"Param: {col.title()}"
            param_rename_dict[col] = new_name
        
        if param_rename_dict:
            self.df.rename(columns=param_rename_dict, inplace=True)
            print(f"Parameter columns renamed: {param_rename_dict}")
        # Build lại cột Strategy
        param_cols_in_df = [col for col in self.df.columns if col.startswith("Param: ")]
        cols_for_strategy = []

        if 'Frequency' in self.df.columns:
            cols_for_strategy.append(self.df['Frequency'].astype(str))
        if 'Inertia' in self.df.columns:
            cols_for_strategy.append(self.df['Inertia'].astype(str))
        if "Threshold" in self.df.columns:
            cols_for_strategy.append(self.df['Threshold'].astype(str))
        if 'Halflife' in self.df.columns:
            cols_for_strategy.append(self.df['Halflife'].astype(str))
        if 'Upper' in self.df.columns:
            cols_for_strategy.append(self.df['Upper'].astype(str))
        if 'Lower' in self.df.columns:
            cols_for_strategy.append(self.df['Lower'].astype(str))
        if 'Velocity' in self.df.columns:
            cols_for_strategy.append(self.df['Velocity'].astype(str))
        if 'Score' in self.df.columns:
            cols_for_strategy.append(self.df['Score'].astype(str))
        if 'Entry' in self.df.columns:
            cols_for_strategy.append(self.df['Entry'].astype(str))
        if 'Exit' in self.df.columns:
            cols_for_strategy.append(self.df['Exit'].astype(str))
        if 'Smooth' in self.df.columns:
            cols_for_strategy.append(self.df['Exit'].astype(str))   
        for col in param_cols_in_df:
            cols_for_strategy.append(self.df[col].astype(str))

        if cols_for_strategy:
            self.df['Strategy'] = cols_for_strategy[0]
            for col_series in cols_for_strategy[1:]:
                self.df['Strategy'] = self.df['Strategy'] + "_" + col_series

        # Xử lý parameters column cũ nếu vẫn còn
        if 'params' in self.df.columns:
            self.df['params'] = self.df['params'].apply(
                lambda x: str(x) if isinstance(x, dict) else str(x)
            )
        
        # Convert percentage columns
        percentage_columns = [
            'Total Return (%)', 'Annual Return (%)',
            'Volatility (%)', 'Win Rate (%)', 'Avg Trade Return (%)',
            'Best Trade (%)', 'Worst Trade (%)',
            
        ]
        
        for col in percentage_columns:
            if col in self.df.columns:
                # Chuyển từ decimal sang percentage (0.1 -> 10%)
                self.df[col] = self.df[col] * 100

        if 'PPC' in self.df.columns:
            # PPC format as currency/decimal
            self.df['PPC'] = self.df['PPC'].round(4)

        # Build sort columns list
        sort_columns = []
        sort_ascending = []
        
        # Primary sort columns
        primary_sorts = [
            ('Alpha Strategy', True),
            ('Inertia', True), 
            ('Threshold', True),
            ('Halflife', True),
            ('Upper', True),
            ('Lower', True),
            ('Velocity', True),
            ('Score', True),
            ('Entry', True),
            ('Exit', True),
            ('Frequency', True),
            ('Smooth', True),
            
        ]
        
        for col_name, ascending in primary_sorts:
            if col_name in self.df.columns:
                sort_columns.append(col_name)
                sort_ascending.append(ascending)
        
        # Add parameter columns to sort
        param_cols_in_df = [col for col in self.df.columns if col.startswith('Param: ')]
        param_cols_in_df.sort()  # Sort parameter names alphabetically
        
        for param_col in param_cols_in_df:
            sort_columns.append(param_col)
            sort_ascending.append(True)  # Ascending for parameters
        
        print(f"Sort order: {sort_columns}")
        
        if sort_columns:
            # Handle mixed data types in parameter columns
            for col in param_cols_in_df:
                if col in self.df.columns:
                    # Convert to string for consistent sorting, then back to numeric if possible
                    try:
                        # Try to convert to numeric first
                        self.df[col] = pd.to_numeric(self.df[col], errors='ignore')
                    except:
                        pass
            
            self.df = self.df.sort_values(sort_columns, ascending=sort_ascending)
            self.df.reset_index(drop=True, inplace=True)
        
        print(f"Data formatted and sorted by: {sort_columns}")
        return True
    
    def apply_excel_formatting(self, worksheet):
        """
        Áp dụng formatting cho Excel worksheet
        """
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply header formatting
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data formatting
        data_alignment = Alignment(horizontal="center", vertical="center")
        
        # Apply formatting to all data cells
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row,
                                     min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.alignment = data_alignment
                cell.border = thin_border
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Max width 50
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Apply conditional formatting cho các metrics quan trọng
        metrics_to_format = [
            ('Sharpe Ratio', 'sharpe'),
            ('MDD (%)', 'mddPct'),
            ('PPC', 'ppc'),
            ('TVR', 'tvr')
        ]
        
        for metric_name, metric_type in metrics_to_format:
            metric_col = None
            for idx, cell in enumerate(worksheet[1], 1):
                if cell.value == metric_name:
                    metric_col = idx
                    break
            
            if metric_col:
                col_letter = worksheet.cell(row=1, column=metric_col).column_letter
                range_str = f"{col_letter}2:{col_letter}{worksheet.max_row}"
                
                if metric_type == 'sharpe':
                    # Sharpe Ratio: red (low) -> yellow (medium) -> green (high)
                    color_scale_rule = ColorScaleRule(
                        start_type='num', start_value=0, start_color='F8696B',
                        mid_type='num', mid_value=2, mid_color='FFEB9C',
                        end_type='num', end_value=5, end_color='63BE7B'
                    )
                elif metric_type == 'mddPct':
                    # Max Drawdown: green (low) -> yellow (medium) -> red (high)
                    color_scale_rule = ColorScaleRule(
                        start_type='num', start_value=0, start_color='63BE7B',
                        mid_type='num', mid_value=50, mid_color='FFEB9C',
                        end_type='num', end_value=100, end_color='F8696B'
                    )
                elif metric_type == 'ppc':
                    # PPC: red (low/negative) -> yellow (neutral) -> green (high)
                    color_scale_rule = ColorScaleRule(
                        start_type='min', start_color='F8696B',
                        mid_type='num', mid_value=0, mid_color='FFEB9C',
                        end_type='max', end_color='63BE7B'
                    )
                elif metric_type == 'tvr':
                    # TVR: green (low) -> yellow (medium) -> red (high)
                    color_scale_rule = ColorScaleRule(
                        start_type='min', start_color='63BE7B',
                        mid_type='percentile', mid_value=50, mid_color='FFEB9C',
                        end_type='max', end_color='F8696B'
                    )
                
                worksheet.conditional_formatting.add(range_str, color_scale_rule)
    
    def export_to_xlsx(self, output_path=None):
        """
        Export data ra file XLSX với formatting
        
        Args:
            output_path (str): Đường dẫn output file. Nếu None thì auto-generate
        """
        if self.df is None:
            print("No data to export. Please run load_data() and format_dataframe() first.")
            return False
        
        if output_path is None:
            base_name = os.path.splitext(os.path.basename(self.pkl_file_path))[0]
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"{base_name}_formatted_{timestamp}.xlsx"
        
        try:
            print(f"Exporting to {output_path}")
            
            # Tạo workbook và worksheet
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Backtest Results"
            
            # Write data
            for r in dataframe_to_rows(self.df, index=False, header=True):
                worksheet.append(r)
            
            # Apply formatting
            self.apply_excel_formatting(worksheet)
            
            # Save file
            workbook.save(output_path)
            print(f"Successfully exported {len(self.df)} records to {output_path}")
            
            # Print preview
            self.print_data_preview()
            
            return True
            
        except Exception as e:
            print(f"Error exporting to XLSX: {str(e)}")
            return False
    
    def print_data_preview(self):
        """
        In preview của data
        """
        if self.df is None:
            return
        
        print(f"\n{'='*80}")
        print("DATA PREVIEW")
        print(f"{'='*80}")
        print(f"Shape: {self.df.shape}")
        print(f"Columns: {list(self.df.columns)}")
        print(f"\nFirst 5 rows:")
        print(self.df.head().to_string())
        
        # Key metrics statistics
        key_metrics = ['Sharpe Ratio', 'MDD (%)', 'PPC', 'TVR', "Net Profit"]
        for metric in key_metrics:
            if metric in self.df.columns:
                print(f"\n{metric} Statistics:")
                print(self.df[metric].describe())
        
        # Parameter columns summary
        param_cols = [col for col in self.df.columns if col.startswith('Param: ')]
        if param_cols:
            print(f"\nParameter Columns Found: {len(param_cols)}")
            for param_col in sorted(param_cols):
                print(f"  - {param_col}: {self.df[param_col].nunique()} unique values")
                if self.df[param_col].dtype in ['int64', 'float64']:
                    print(f"    Range: {self.df[param_col].min()} - {self.df[param_col].max()}")
                else:
                    top_values = self.df[param_col].value_counts().head(3)
                    print(f"    Top values: {dict(top_values)}")


class ScanParams:
    """
    Ultra-optimized generator for alpha parameter scans.
    """

    def __init__(self, lst_alpha_names, alpha_params, min_freq, max_freq, step_freq, fee, gen=None):
        self.alpha_params = alpha_params or {}
        self.min_freq, self.max_freq, self.step_freq  = min_freq, max_freq, step_freq
        self.fee, self.gen = fee, gen
        self.lst_reports = []

        if lst_alpha_names:
            self.lst_reports = self.gen_lst_reports(lst_alpha_names)

    # ---------- Param utilities ----------
    @staticmethod
    def gen_list():
        return [round(i * 0.1, 1) for i in range(1, 10)]

    def gen_band_list(self):
        vals = self.gen_list()
        return [(x, y) for x, y in product(vals + [1.0], vals) if x > y]
    
    def gen_smooth_list(self):
        vals = self.gen_list()
        return [(x, y) for x, y in product(vals, vals) if x > y]
    
    def gen_score_list(self):
        scores = [3, 4, 5, 6, 7, 8]
        entries = [1, 2, 3, 4]
        exits = [0, 1, 2]
        return [(s, e1, e2) for s, e1, e2 in product(scores, entries, exits) if e1 > e2]
        
    def gen_params_combinations(self):
        names, ranges = [], []
        for k, v in self.alpha_params.items():
            if isinstance(v, dict):
                values = np.arange(v["start"], v["end"] + v["step"], v["step"])
            elif isinstance(v, list):
                values = v
            else:
                values = [v]
            names.append(k)
            ranges.append(values)
        return [dict(zip(names, c)) for c in product(*ranges)]

    # ---------- Core generator ----------
    def gen_lst_reports(self, lst_alpha_names):
        freqs = range(self.min_freq, self.max_freq + self.step_freq, self.step_freq)
        params = self.gen_params_combinations()
        gl = self.gen_list

        # map structure: gen -> (extra_fields, value_generators)
        map_gen = {
            "1_1": (["threshold", "halflife"], [gl(), [0.0]+gl()]), 
            "1_2": (["upper_lower"], [self.gen_band_list()]),
            "1_3": (["score_entry_exit"], [self.gen_score_list()]),
            "1_4": (["entry_exit", "smooth"], [self.gen_smooth_list(), [1,2,3,4]]),
        }

        if self.gen not in map_gen:
            raise ValueError(f"Unsupported gen mode: {self.gen}")

        keys, values = map_gen[self.gen]
        combos = product(lst_alpha_names, freqs, *values, params)
        fee = self.fee
        gen = self.gen

        # Use list comprehension + tuple unpacking for speed
        reports = []
        append = reports.append
        for c in combos:
            base = {"alphaName": c[0], "freq": c[1], "fee": fee, "params": c[-1]}

            if gen == "1_1":
                base.update({"threshold": c[2], "halflife": c[3]})
            elif gen == "1_2":
                upper, lower = c[2]
                base.update({"upper": upper, "lower": lower})
            elif gen == "1_3":
                score, entry, exit = c[2]
                base.update({"score": score, "entry": entry, "exit": exit})
            elif gen == "1_4":
                entry, exit = c[2]
                base.update({"entry": entry, "exit": exit, "smooth": c[3]})

            append(base)

        return reports



def load_dic_freqs(source):
    if source == "volume_bar":
        fn = "/home/ubuntu/nevir/gen/dic_freqs_volume_bar.pickle"
    elif source == "dollar_bar":
        fn = "/home/ubuntu/nevir/gen/dic_freqs_dollar_bar.pickle"
    elif source == "ha":
        fn = "/home/ubuntu/nevir/gen/dic_freqs_ha.pickle"
    elif source == "ha_confirm":
        fn = "/home/ubuntu/nevir/gen/dic_freqs_ha_confirm.pickle"
    else:
        fn = "/home/ubuntu/nevir/gen/alpha.pkl"
    with open(fn, 'rb') as file:
        DIC_FREQS = pickle.load(file)

    return DIC_FREQS


def run_single_backtest(config, dic_freqs, DIC_ALPHAS,df_tick=None,gen=None,start=None,end=None):
    """
    Chạy backtest cho một config duy nhất
    """
    gen_params = {}
    if gen == "1_3":
        gen_params['score'] = config['score']
        gen_params['entry'] = config['entry']
        gen_params['exit'] = config['exit']
    elif gen == "1_2":
        gen_params['upper'] = config['upper']
        gen_params['lower'] = config['lower']
    elif gen == "1_1":
        gen_params['threshold'] = config['threshold']
        gen_params['halflife'] = config['halflife']
    elif gen == "1_4":
        gen_params['entry'] = config['entry']
        gen_params['exit'] = config['exit']
        gen_params['smooth'] = config['smooth']

    bt = Simulator(
        alpha_name=config['alphaName'],
        freq=config['freq'],
        gen_params=gen_params,
        fee=config['fee'],
        df_alpha=dic_freqs[config['freq']].copy(),
        params=config.get('params', {}),
        DIC_ALPHAS=DIC_ALPHAS,
        df_tick=None,
        gen=gen,
        start=start,
        end=end
    )
    bt.compute_signal()
    bt.compute_position()
    # bt.change_cut_time(dic_freqs[1].copy(), "14:25:00")
    bt.compute_tvr_and_fee()
    bt.compute_profits()
    bt.compute_performance(start=start, end=end)
    
    report_with_params = bt.report.copy()
    

    
    params = config.get('params', {})
    for param_name, param_value in params.items():
        report_with_params[f"param_{param_name}"] = param_value
    # print(f"Report", report_with_params['sharpe'])
    return report_with_params
    


def run_backtest_batch(configs_batch, dic_freqs, DIC_ALPHAS, df_tick=None,start=None,end=None,gen=None):
    """
    Chạy backtest cho một batch configs (để tối ưu overhead của multiprocessing)
    """
    batch_results = []
    for config in configs_batch:
        result = run_single_backtest(config, dic_freqs, DIC_ALPHAS, df_tick=df_tick,start=start,end=end,gen=gen)
        if result is not None:
            batch_results.append(result)
            
    return batch_results


def split_list_into_batches(lst, batch_size):
    """
    Chia list thành các batches nhỏ hơn
    """
    for i in range(0, len(lst), batch_size):
        yield lst[i:i + batch_size]


class MultiThreadScanner:
    def __init__(self, n_workers=None, use_processes=True, batch_size=10):
        """
        n_workers: số worker threads/processes. Nếu None thì dùng CPU count
        use_processes: True để dùng ProcessPoolExecutor, False để dùng ThreadPoolExecutor
        batch_size: số configs để group lại thành 1 batch (giảm overhead)
        """
        self.n_workers = n_workers or cpu_count()
        self.use_processes = use_processes
        self.batch_size = batch_size
        self.executor_class = ProcessPoolExecutor if use_processes else ThreadPoolExecutor

    def run_parallel_backtest(self, lst_configs, dic_freqs, DIC_ALPHAS, df_tick=None,gen=None,start=None,end=None):
        """
        Chạy backtest song song cho tất cả configs
        """
        print(f"Running parallel backtest with {self.n_workers} {'processes' if self.use_processes else 'threads'}")
        print(f"Total configs: {len(lst_configs)}")
        
        # Chia configs thành batches để giảm overhead
        config_batches = list(split_list_into_batches(lst_configs, self.batch_size))
        print(f"Split into {len(config_batches)} batches of size {self.batch_size}")
        
        lst_passed = []
        completed_batches = 0
        total_configs_processed = 0
        
        start_time = time.time()
        batch_times = []  # Lưu thời gian của từng batch
        
        with self.executor_class(max_workers=self.n_workers) as executor:
            # Submit tất cả batches với thời gian bắt đầu
            future_to_batch_info = {}
            for i, batch in enumerate(config_batches):
                future = executor.submit(run_backtest_batch, batch, dic_freqs, DIC_ALPHAS,df_tick=df_tick,gen=gen,start=start,end=end)
                future_to_batch_info[future] = {
                    'batch_idx': i,
                    'batch_size': len(batch),
                    'submit_time': time.time()
                }
            
            # Collect results khi hoàn thành
            for future in as_completed(future_to_batch_info):
                batch_info = future_to_batch_info[future]
                batch_idx = batch_info['batch_idx']
                batch_size = batch_info['batch_size']
                batch_start_time = batch_info['submit_time']
                try:
                    batch_completion_time = time.time()
                    batch_duration = batch_completion_time - batch_start_time
                    batch_times.append(batch_duration)
                    
                    batch_results = future.result()
                    
                    if batch_results:
                        lst_passed.extend(batch_results)
                    
                    completed_batches += 1
                    total_configs_processed += batch_size
                    progress = completed_batches / len(config_batches) * 100
                    
                    # Tính toán thống kê tốc độ
                    configs_per_second = batch_size / batch_duration
                    avg_batch_time = sum(batch_times) / len(batch_times)
                    
                    # Ước tính thời gian còn lại
                    remaining_batches = len(config_batches) - completed_batches
                    estimated_remaining_time = remaining_batches * avg_batch_time
                    
                    print(f"Batch {batch_idx + 1:3d}/{len(config_batches)} completed in {batch_duration:5.1f}s "
                            f"({configs_per_second:4.1f} configs/s) - "
                            f"Found {len(batch_results):2d}/{batch_size} passed - "
                            f"Progress: {total_configs_processed:4d}/{len(lst_configs)} configs ({progress:5.1f}%) - "
                            f"Total passed: {len(lst_passed):3d} - "
                            f"ETA: {estimated_remaining_time:5.1f}s")
                except Exception as e:
                    print(f"Error in batch {batch_idx}: {str(e)}")
            
        
        total_time = time.time() - start_time
        
        # Thống kê chi tiết
        print(f"\n{'='*80}")
        print(f"PERFORMANCE SUMMARY")
        print(f"{'='*80}")
        print(f"Total execution time: {total_time:.1f} seconds ({total_time/60:.1f} minutes)")
        print(f"Total configs processed: {total_configs_processed:,}")
        print(f"Total configs passed: {len(lst_passed):,}")
        print(f"Success rate: {len(lst_passed)/total_configs_processed*100:.2f}%")
        print(f"Average processing speed: {total_configs_processed/total_time:.1f} configs/second")
        print(f"Average batch time: {sum(batch_times)/len(batch_times):.2f}s")
        print(f"Fastest batch: {min(batch_times):.2f}s")
        print(f"Slowest batch: {max(batch_times):.2f}s")
        print(f"Batch time std dev: {np.std(batch_times):.2f}s")
        
        if self.use_processes:
            theoretical_speedup = self.n_workers
            actual_speedup = total_configs_processed / total_time / (total_configs_processed / total_time / self.n_workers)
            efficiency = actual_speedup / theoretical_speedup * 100
            print(f"Theoretical speedup: {theoretical_speedup}x")
            print(f"Parallel efficiency: {efficiency:.1f}%")
        
        return lst_passed

def get_drive_service():
    SCOPES = ['https://www.googleapis.com/auth/drive']
    CREDENTIAL_FILE = "/home/ubuntu/nevir/credentials.json"
    creds = None
    if os.path.exists('/home/ubuntu/nevir/token.json'):
        creds = Credentials.from_authorized_user_file('/home/ubuntu/nevir/token.json', SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIAL_FILE, SCOPES)
            creds = flow.run_local_server(port=0)
        with open('/home/ubuntu/nevir/token.json', 'w') as token:
            token.write(creds.to_json())
    return build('drive', 'v3', credentials=creds)


def upload_xlsx_to_gdrive(
    save_file: str,
    parent_folder_id: str,
    output_json: str,
    upload_results: dict,
    result_key: str,
    max_retries: int = 5,
    base_delay: int = 5
):
    """
    Upload XLSX -> Google Drive (convert to Google Sheet)
    Retry an toàn nếu timeout / lỗi mạng

    Thành công khi:
    - upload OK
    - share public OK
    - lưu link vào json OK
    """

    if not os.path.exists(save_file):
        raise FileNotFoundError(f"File not found: {save_file}")

    sheet_name = os.path.splitext(os.path.basename(save_file))[0]
    service = get_drive_service()

    file_metadata = {
        "name": sheet_name,
        "mimeType": "application/vnd.google-apps.spreadsheet",
        "parents": [parent_folder_id],
    }

    for attempt in range(1, max_retries + 1):
        file_id = None
        try:
            print(f"📤 Upload attempt {attempt}/{max_retries}: {sheet_name}")

            # ⚠️ MUST recreate media every retry
            media = MediaFileUpload(
                save_file,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                resumable=True,
            )

            # ---------- Upload ----------
            response = service.files().create(
                body=file_metadata,
                media_body=media,
                fields="id",
            ).execute()

            file_id = response.get("id")
            if not file_id:
                raise RuntimeError("Upload succeeded but no file_id returned")

            # ---------- Share public ----------
            permission = {
                "type": "anyone",
                "role": "reader",
            }

            service.permissions().create(
                fileId=file_id,
                body=permission,
                fields="id",
            ).execute()

            # ---------- Save result ----------
            sheet_url = f"https://docs.google.com/spreadsheets/d/{file_id}"

            upload_results[result_key] = sheet_url
            with open(output_json, "w", encoding="utf-8") as f:
                json.dump(upload_results, f, indent=2, ensure_ascii=False)

            print(f"✅ Upload & share OK: {sheet_url}")
            print(f"🔗 Updated JSON: {output_json}")

            return sheet_url

        except HttpError as e:
            print(f"⚠️ HttpError (attempt {attempt}): {e}")

        except Exception as e:
            print(f"⚠️ Error (attempt {attempt}): {e}")

        # ---------- Retry ----------
        if attempt < max_retries:
            sleep_time = base_delay * attempt
            print(f"🔁 Retry after {sleep_time}s...\n")
            time.sleep(sleep_time)
        else:
            raise RuntimeError(
                f"❌ Upload failed after {max_retries} attempts: {sheet_name}"
            )
            
def generate_time_ranges(mode):
    if mode == "FULL":
        return [{
            "label": "2018_2025",
            "start": "2018_01_01",
            "end": "2025_09_08"
        }]

    if mode == "YEARLY":
        return [
            {
                "label": str(y),
                "start": f"{y}_01_01",
                "end": f"{y}_12_31"
            } for y in range(2019, 2025)
        ]

    if mode == "WFO":
        segments = [
            ("2018_01_01", "2020_01_01"),
            ("2018_07_01", "2020_07_01"),
            ("2019_01_01", "2021_01_01"),
            ("2019_07_01", "2021_07_01"),
            ("2020_01_01", "2022_01_01"),
            ("2020_07_01", "2022_07_01"),
            ("2021_01_01", "2023_01_01"),
            ("2021_07_01", "2023_07_01"),
            ("2022_01_01", "2024_01_01"),
            ("2022_07_01", "2024_07_01"),
            ("2023_01_01", "2025_01_01"),
            ("2023_07_01", "2025_07_01"),
            ("2024_01_01", "2026_01_01"),
        ]
        return [
            {
                "label": f"{s}_{e}",
                "start": s,
                "end": e
            } for s, e in segments
        ]

    raise ValueError(f"Unknown mode: {mode}")

if __name__ == "__main__":
    MODE = "FULL"          # ← đổi duy nhất dòng này
    gen = "1_1"
    source = None
    start_time = time.time()
    folder_path = f"/home/ubuntu/nevir/gen/results_{gen}/"
    os.makedirs(folder_path, exist_ok=True)
    OUTPUT_JSON = f"/home/ubuntu/nevir/gen/r_{gen}.json"
    if os.path.exists(OUTPUT_JSON):
        with open(OUTPUT_JSON, 'r') as f:
            upload_results = json.load(f)
    else:
        upload_results = {}
    time_ranges = generate_time_ranges(MODE)
    alpha_list = [
        # "alpha_full_factor_062",
        # "alpha_full_factor_095"
        
        # "alpha_full_factor_100",
        # "alpha_full_factor_101"
        
        # "alpha_full_factor_066",
        
        # "alpha_full_factor_080",
        # "alpha_full_factor_072",
        # "alpha_full_factor_090",
        # "alpha_full_factor_092",
        
        # "alpha_full_factor_093",
        # "alpha_full_factor_098"
        
        # "alpha_full_factor_099",
        
        # "alpha_full_factor_100",
        # "alpha_full_factor_101"

        #huy
        # 'alpha_full_factor_007',
        # 'alpha_full_factor_031',
        # 'alpha_full_factor_034',
        # 'alpha_full_factor_046', #window 5-200-5
        # 'alpha_full_factor_066', #window 10-100-10 factor 1-4-1
        # 'alpha_full_factor_080', #window 10-100-10 factor 10-40-10
        # 'alpha_full_factor_099', #fast 6-24-6 slow 24-84-6
        # 'alpha_full_factor_100', #window_short 12-24-6 window_long 24-180-12 
        # 'alpha_full_factor_101', #window_short 12-24-6 window_long 24-180-12 
        # 'alpha_full_factor_105' ##window 5-200-5

        # 'alpha_full_factor_001'


        # 'alpha_full_factor_062_zscore_clipping'
        # 'alpha_full_factor_095_regime_adaptive'
        'alpha_full_factor_046_vol_weighted' 
        

    ]
    for alpha_name in alpha_list:
        alpha_params = {
            
            # "window_reg": {"start": 5, "end": 100, "step": 10 },
            # "window_signal": {"start": 5, "end": 20, "step": 5},
            # "window_delta": {"start": 1, "end": 4, "step": 1},
            "window": {"start": 5, "end": 200, "step": 5 },
            # "factor": {"start": 10, "end": 40, "step": 10},
            # "window_long": {"start": 24, "end": 180, "step": 12},

            # "fast": {"start": 6, "end": 24, "step": 6},
            # "slow": {"start": 24, "end": 84, "step": 6},
            #  "factor": {"start": 10, "end": 40, "step": 10},
            # "short_delta": {"start": 1, "end": 4, "step": 1},
            # "long_delta":{"start":4, "end":7, "step":1},
            # "window_rank":{"start":10, "end":30 , "step":10},
        }

        dic_freqs = load_dic_freqs(source)
        DIC_ALPHAS = Domains.get_list_of_alphas()

        for tr in time_ranges:
            start = tr["start"]
            end = tr["end"]
            year_label = tr["label"]

            save_file = os.path.join(
                folder_path,
                f"{alpha_name}_{year_label}.xlsx"
            )

            print(f"=== {MODE} | {alpha_name} | {year_label} ===")

            scan_params = ScanParams(
                lst_alpha_names=[alpha_name],
                alpha_params=alpha_params,
                min_freq=10,
                max_freq=100,
                step_freq=1,
                fee=0.175,
                gen=gen
            )
            lst_configs = scan_params.lst_reports

            scanner_process = MultiThreadScanner(
                n_workers=40, 
                use_processes=True,
                batch_size=1000 
            )


            lst_passed_parallel = scanner_process.run_parallel_backtest(
                lst_configs, dic_freqs, DIC_ALPHAS,
                df_tick=None, start=start, end=end,gen=gen
            )

            if lst_passed_parallel:
                df = pd.DataFrame(lst_passed_parallel)
                exporter = PKLToXLSXExporter(df,gen)
                exporter.format_dataframe()
                exporter.export_to_xlsx(save_file)
                print(f"✅ Saved results to {save_file}")
                try:
                    upload_xlsx_to_gdrive(
                        save_file=save_file,
                        parent_folder_id="1ZI65HWxDFaPcXQYyJFdu47cLQtwJ_4Iw",
                        output_json=OUTPUT_JSON,
                        upload_results=upload_results,
                        result_key=f"{alpha_name}_{year_label}",
                        max_retries=3,
                        base_delay=5
                    )
                except Exception as e:
                    print(f"   ❌ Upload failed after retries: {e}")
                    continue
            else:
                print(f"❌ No configs passed for {year_label}")

    print(f"\nAll done in {time.time() - start_time:.1f} seconds.")